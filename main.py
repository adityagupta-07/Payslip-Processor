from python_docx_replace import docx_replace
from docx import Document
from docxtpl import DocxTemplate
from datetime import datetime
from docx2pdf import convert
import dxpdf
import calendar
import openpyxl
import shutil
import pymupdf
import copy
import os

def user_input():
    file_path = input("Provide file path: ")
    return {
        "file_path": file_path.replace('"', '')
    }

def load_excel_file(user_input):
    workbook = openpyxl.load_workbook(user_input["file_path"], data_only=True)
    sheet = workbook.active
    return sheet

def matching_row_numbers(search_string, sheet):
    matching_rows = []
    for target_row in sheet.iter_rows():
        for cell in target_row:
            if cell.value == search_string:
                if cell.row not in matching_rows:
                    matching_rows.append(cell.row)
    return matching_rows

def get_employee_block(string1, string2, sheet):
    string1_matching_row_numbers = matching_row_numbers(string1, sheet)
    string2_matching_row_numbers = matching_row_numbers(string2, sheet)
    employee_block = list(zip(string1_matching_row_numbers, string2_matching_row_numbers))
    return employee_block

def delete_contents(folder_paths):
    for folder_path in folder_paths:
        shutil.rmtree(folder_path)
        os.makedirs(folder_path, exist_ok=True)
    return

def empty_nested_dict(dictionary):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            empty_nested_dict(value)
        else:
            dictionary[key] = ""
    return dictionary

def docx_creation(employee, placeholders_with_values):
    month_name = employee["specials"]["Month_year"].strftime("%B")
    year = employee["specials"]["Month_year"].strftime("%Y")
    if employee["Employee ID"] == "":
        template_file = "Templates/Template_no_id.docx"
        file_name = (f"{employee["Employee Name"].strip().replace(" ", "_")}_{month_name}_{year}")
    else:
        template_file = "Templates/Template_id.docx"
        file_name = (f"{employee["Employee Name"].strip().replace(" ", "_")}_{month_name}_{year}_{str(employee['Employee ID']).strip()}")
    destination_file = f"Tmp/{file_name}.docx" 
    shutil.copy2(template_file, destination_file)  
    doc = DocxTemplate(destination_file) 
    doc.render(placeholders_with_values)
    doc.save(destination_file)
    return

def get_details_per_employee(sheet, block, data_dict):
    found_ssf_contribution_by_employer = False
    for r in range(block[0], block[1] + 1): # r loops from 1 to 20 (exclusive)
        for target_cell in sheet[r]: # target_cell loops through all the cells in row(r), means row(1), row(2), row(3), ...
            target_value = target_cell.value.strip() if isinstance(target_cell.value, str) else target_cell.value            
            if target_value in data_dict:
                next_cell = sheet.cell(row=r, column=(target_cell.column+1))
                if found_ssf_contribution_by_employer is True and "SSF Contribution by Employer" in target_value:
                    data_dict["duplicates"][target_value] = f"Rs. {next_cell.value:,.2f}"
                    break
                if "SSF Contribution by Employer" in target_value:
                    found_ssf_contribution_by_employer = True
                    data_dict["specials"]["Financial_Year_Note"] = sheet.cell(row=r, column=(target_cell.column+2)).value
                if isinstance(next_cell.value, (int, float)) and (r > block[1]-13 and r < block[1]+1):
                    data_dict[target_value] = f"Rs. {next_cell.value:,.2f}"
                else:
                    data_dict[target_value] = next_cell.value
                if r == (block[1]-2) and target_cell.column == 3:
                    # next_cell.value is returning <class 'datetime.datetime'> (2026-05-01 00:00:00) so we can change the format (%B = August, %Y = 2026)
                    data_dict["specials"]["Month_year"] = next_cell.value
                    data_dict[target_value] = next_cell.value.strftime("%b %Y")
    return data_dict

def batch_convert_docx_to_pdf(input_dir, output_dir):
    # MS Word Independent (but messes up the format)
    for filename in os.listdir(input_dir):
        if filename.endswith(".docx") and not filename.startswith("~$"):
            docx_path = os.path.join(input_dir, filename)
            pdf_filename = filename.rsplit(".", 1)[0] + ".pdf"
            pdf_path = os.path.join(output_dir, pdf_filename)
            try:
                with open(docx_path, "rb") as f:
                    docx_bytes = f.read()
                pdf_bytes = dxpdf.convert(docx_bytes)
                with open(pdf_path, "wb") as f:
                    f.write(pdf_bytes)
            except Exception as e:
                print(f"Failed to convert {filename}. Error: {e}")
    return

def batch_convert_docx_to_pdf1(input_dir, output_dir):
    # MS Word Dependent (Preserves the format)
    convert(input_dir, output_dir)
    return

def master_pdf_creation(destination_folder, month, year):
    pdf_name_list = os.listdir(destination_folder)
    result = pymupdf.open()
    for pdf in pdf_name_list:
        with pymupdf.open(f"./PDFs/{pdf}") as mfile:
            result.insert_pdf(mfile)
    output_dir = "./Master Pdf"
    output_path = f"{output_dir}/Payslip - {month} {year}.pdf"
    result.save(output_path)
    result.close()
    return

def fill_placeholders_in_docx(employee):
    return {
        "MONTH_YEAR_UPPER": f"{employee["specials"]["Month_year"].strftime("%B").upper()} {employee["specials"]["Month_year"].strftime("%Y")}",
        "EMPLOYEE_ID": employee["Employee ID"], 
        "EMPLOYEE_NAME": employee["Employee Name"], 
        "DESIGNATION": employee["Designation"], 
        "PAN": employee["PAN"],
        "CONTACT_NUMBER": employee["Contact Number"], 
        "BASIC_SALARY": employee["Basic Salary"], 
        "ALLOWANCES": employee["Allowances"], 
        "GROSS_SALARY": employee["Gross Salary"], 
        "GROSS_SALARY_WORKING_HOURS": employee["Gross Salary as per working hours"], 
        "SSF_EMPLOYER": employee["SSF Contribution by Employer"], 
        "BONUS": employee["Bonus"], 
        "TOTAL": employee["Total"], 
        "SSF_EMPLOYER1": employee["duplicates"]["SSF Contribution by Employer"], 
        "SSF_EMPLOYEE": employee["SSF Contribution by Employee"], 
        "TDS_FOR_MONTH": employee["TDS for the month"], 
        "TOTAL_DEDUCTION": employee["Total Deduction"], 
        "NET_SALARY_PAID": employee["Net Salary Paid"], 
        "ACCOUNT_NUMBER": employee["Account Number"], 
        "BANK_NAME": employee["Bank Name"], 
        "BRANCH": employee["Branch"], 
        "MARITAL_STATUS": employee["Marital Status"], 
        "ANNUAL_TAXABLE_SALARY": employee["Annual Taxable Salary"], 
        "ANNUAL_SSF_DEPOSIT": employee["Annual SSF deposit"], 
        "ANNUAL_TDS_PAYMENT": employee["Annual TDS Payment"], 
        "ANNUAL_NET_SALARY": employee["Annual Net Salary"], 
        "FINANCIAL_YEAR_NOTE": employee["specials"]["Financial_Year_Note"], 
        "MONTH_YEAR": employee["Month"]
    }

employee_dictionary = {
    "Employee ID": "",
    "Employee Name": "",
    "Designation": "",
    "PAN": "",
    "Contact Number": "",
    "Basic Salary": "",
    "Allowances": "",
    "Gross Salary": "",
    "Gross Salary as per working hours": "",
    "SSF Contribution by Employer": "",
    "Bonus": "",
    "Total": "",
    "SSF Contribution by Employee": "", 
    "TDS for the month": "", 
    "Total Deduction": "", 
    "Net Salary Paid": "", 
    "Account Number": "", 
    "Bank Name": "",
    "Branch": "",
    "Marital Status": "",
    "Annual Taxable Salary": "",
    "Annual SSF deposit": "",
    "Annual TDS Payment": "",
    "Annual Net Salary": "",
    "Month": "",
    "duplicates": {
        "SSF Contribution by Employer": ""
    },
    "specials": {
        "Financial_Year_Note": "",
        "Month_year": ""
    }
}





input_excel_file_path = user_input()

sheet = load_excel_file(input_excel_file_path)

employee_blocks = get_employee_block("EMPLOYEE INFORMATION", "Net Salary Paid", sheet)

delete_contents(["./Tmp", "./PDFs", "./Master Pdf"])

employee_details = [] # stores dictionary per employee in a list

for block in employee_blocks: #[(1, 19), (22, 41), ......]
    employee_dictionary = empty_nested_dict(employee_dictionary)
    employee_dictionary = get_details_per_employee(sheet, block, employee_dictionary)
    employee_details.append(copy.deepcopy(employee_dictionary)) # Without copy.deepcopy(), it appends reference to the same underlying dictionary every time. With copy.deepcopy(), it appends a fresh original dictionary, not its reference or instance.

for employee in employee_details:
    placeholders_in_docx_with_values = fill_placeholders_in_docx(employee)
    docx_creation(employee, placeholders_in_docx_with_values)

docs_folder = "./Tmp"
destination_folder = "./PDFs"

# batch_convert_docx_to_pdf(docs_folder, destination_folder) # MS Word Independent
batch_convert_docx_to_pdf1(docs_folder, destination_folder) # MS Word Dependent

master_pdf_creation(destination_folder, employee_details[0]["specials"]["Month_year"].strftime("%B"), employee_details[0]["specials"]["Month_year"].strftime("%Y"))